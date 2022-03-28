import os.path
import datetime
from time import strptime

import openpyxl

dt_now = datetime.datetime.now()
dir_path = "D:/lifeRecord"
file_path = f"{dir_path}/{dt_now.date()}.xlsx"

def save_excel(task, category, startTime):
    if category == "" or category == None:
        return
    if os.path.exists(dir_path) == False:
        os.mkdir(dir_path)
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1, value='시작시간')
        sheet.cell(row=1, column=2, value='총 시간')
        sheet.cell(row=1, column=3, value='카테고리')
        sheet.cell(row=1, column=4, value='상세')

    if sheet.max_row != 1:
        update_total_time(sheet, sheet.max_row, startTime)
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1, value=startTime)
    sheet.cell(row=last_row, column=3, value=category)
    sheet.cell(row=last_row, column=4, value=task)

    wb.save(file_path)

def update_total_time(sheet, done_row, endTime):
    start_Time = strptime(sheet.cell(row=done_row,column=1).value,"%H:%M:%S")
    end_time = strptime(endTime,"%H:%M:%S")
    start = datetime.timedelta(hours=start_Time.tm_hour,minutes=start_Time.tm_min,seconds=start_Time.tm_sec)
    end = datetime.timedelta(hours=end_time.tm_hour,minutes=end_time.tm_min,seconds=end_time.tm_sec)
    totalTime =  end-start
    sheet.cell(row=done_row,column=2,value=totalTime)

def closing_a_day():
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    category = {}
    for i in range(sheet.max_row):
        if i == 0:
            continue
        cur_category = sheet.cell(row=i+1, column=3).value
        cur_hour = sheet.cell(row=i+1, column=2).value
        if cur_hour == None:
            curtime = datetime.datetime.now().time()
            currentTime = f'{curtime.hour}:{curtime.minute}:{curtime.second}'
            end_time = strptime(currentTime,"%H:%M:%S")
            end = datetime.timedelta(hours=end_time.tm_hour, minutes=end_time.tm_min, seconds=end_time.tm_sec)
            start_Time = strptime(sheet.cell(row=i+1, column=1).value, "%H:%M:%S")
            start = datetime.timedelta(hours=start_Time.tm_hour, minutes=start_Time.tm_min, seconds=start_Time.tm_sec)
            cur_hour = end - start
            sheet.cell(row=i + 1, column=2, value=cur_hour)
        if cur_category in category:
            category[cur_category] = category[cur_category] + cur_hour
        else:
            category[cur_category] = cur_hour

    sheet.cell(row=1, column=6, value='카테고리')
    sheet.cell(row=1, column=7, value='총 시간')
    for idx, key in enumerate(category):
        sheet.cell(row = 2 + idx, column= 6, value=key)
        sheet.cell(row = 2 + idx, column= 7, value=category[key])

    wb.save(file_path)